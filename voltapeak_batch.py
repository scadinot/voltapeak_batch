"""
voltapeak_batch
===============

Application d'analyse automatisÃ©e de fichiers de voltampÃ©romÃ©trie Ã  onde
carrÃ©e (SWV â *Square Wave Voltammetry*).

Pour chaque fichier ``.txt`` d'un dossier sÃ©lectionnÃ©, le script :
    1. lit les deux colonnes (Potentiel, Courant) ;
    2. nettoie et inverse le signe du courant (convention SWV) ;
    3. lisse le signal par filtre de Savitzky-Golay ;
    4. dÃ©tecte la position approximative du pic ;
    5. estime la ligne de base par l'algorithme asPLS
       (*Asymmetric Penalized Least Squares*) en excluant la zone du pic ;
    6. corrige le signal par soustraction de la baseline ;
    7. sauvegarde un graphique PNG ainsi que, en option, un CSV ou un XLSX
       par fichier ;
    8. agrÃ¨ge les rÃ©sultats multi-Ã©lectrodes (convention de nommage
       ``<base>_C<NN>.txt``) dans un unique fichier Excel rÃ©capitulatif,
       avec injection d'une formule ``= Courant / FrÃ©q`` pour la charge.

Les traitements par fichier sont parallÃ©lisÃ©s avec
``multiprocessing.Pool`` et l'utilisateur pilote l'outil via une interface
graphique Tkinter.

DÃ©pendances externes : ``numpy``, ``pandas``, ``matplotlib``,
``pybaselines``, ``scipy``, ``openpyxl``. Tkinter est inclus dans la
bibliothÃ¨que standard.
"""

import glob
import os
import platform
import re
import subprocess
import time
from multiprocessing import Pool, cpu_count, freeze_support
from tkinter import Button, Frame, IntVar, Label, Radiobutton, StringVar, Text, Tk, filedialog, messagebox, ttk
from typing import cast

import matplotlib
from numpy.typing import NDArray

# Backend "Agg" (non-interactif) : on ne fait que ``savefig`` ; les workers
# multiprocessing n'ont pas de boucle Tk et un backend GUI y serait inutile
# (risque de crash sur plate-forme headless). ``use()`` doit prÃ©cÃ©der ``pyplot``.
matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pybaselines.whittaker import aspls
from scipy.signal import savgol_filter


def open_folder(path):
    """Ouvre un dossier dans l'explorateur de fichiers natif du systÃ¨me.

    L'implÃ©mentation diffÃ¨re selon l'OS : Windows expose ``os.startfile``,
    macOS s'appuie sur ``open`` et Linux sur ``xdg-open``.

    ParamÃ¨tres:
        path (str): chemin absolu du dossier Ã  ouvrir.
    """
    if platform.system() == "Windows":
        os.startfile(path)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", path])
    else:  # Linux
        subprocess.call(["xdg-open", path])

def readFile(filePath, sep, decimal) -> (pd.DataFrame|None):
    """Charge un fichier SWV en DataFrame Ã  deux colonnes.

    La premiÃ¨re ligne est supposÃ©e Ãªtre un entÃªte (mÃ©tadonnÃ©es du
    potentiostat) et est ignorÃ©e. Seules les deux premiÃ¨res colonnes sont
    lues, renommÃ©es respectivement ``Potential`` et ``Current``.

    ParamÃ¨tres:
        filePath (str): chemin du fichier ``.txt`` Ã  lire.
        sep (str): sÃ©parateur de colonnes (``"\\t"``, ``","``, ``";"``
            ou ``" "``).
        decimal (str): sÃ©parateur dÃ©cimal (``"."`` ou ``","``).

    Retourne:
        pandas.DataFrame: DataFrame avec colonnes ``Potential`` et
        ``Current``.

    Notes:
        L'encodage Latin-1 est retenu pour tolÃ©rer les caractÃ¨res
        accentuÃ©s Ã©ventuels dans l'entÃªte mÃ©tier (exports Windows).
    """
    with open(filePath, encoding="latin1") as fileStream:
        dataFrame = pd.read_csv(fileStream, sep=sep, skiprows=1, usecols=[0, 1], names=["Potential", "Current"], decimal=decimal)
    return dataFrame

def processData(dataFrame) -> tuple:
    """Nettoie et prÃ©pare les donnÃ©es pour la chaÃ®ne de traitement.

    Trois Ã©tapes :
        1. suppression des lignes dont le courant est exactement nul
           (artefacts de mesure ou points de saturation) ;
        2. tri par potentiel croissant (indispensable pour asPLS et
           pour le calcul de pente dans la dÃ©tection de pic) ;
        3. inversion du signe du courant : en SWV les pics cathodiques
           sont mesurÃ©s nÃ©gatifs par le potentiostat ; on les ramÃ¨ne
           vers le haut afin que le reste du pipeline manipule toujours
           des pics positifs.

    ParamÃ¨tres:
        dataFrame (pandas.DataFrame): DataFrame brut issu de ``readFile``.

    Retourne:
        tuple:
            - potentialValues (numpy.ndarray): potentiels triÃ©s (V).
            - signalValues (numpy.ndarray): courants inversÃ©s (A).
            - dataFrame (pandas.DataFrame): DataFrame nettoyÃ© (sert de
              base Ã  l'export CSV/XLSX).
    """
    dataFrame = dataFrame[dataFrame["Current"] != 0].sort_values("Potential").reset_index(drop=True)
    potentialValues = dataFrame["Potential"].values
    signalValues = -dataFrame["Current"].values  # Convention SWV : on ramÃ¨ne les pics vers le haut.
    return potentialValues, signalValues, dataFrame

def smoothSignal(signalValues: NDArray[np.float64]) -> NDArray[np.float64]:
    """Lisse le signal par un filtre polynomial de Savitzky-Golay.

    Le filtre prÃ©serve l'amplitude et la position du pic tout en
    attÃ©nuant le bruit haute frÃ©quence. La fenÃªtre de 11 points et
    l'ordre 2 constituent un compromis Ã©prouvÃ© pour les signaux SWV
    typiques.

    ParamÃ¨tres:
        signalValues (numpy.ndarray): signal brut Ã  lisser (A).

    Retourne:
        numpy.ndarray: signal lissÃ©, mÃªme forme que l'entrÃ©e.
    """
    return np.asarray(savgol_filter(signalValues, window_length=11, polyorder=2))

def getPeakValue(signalValues, potentialValues, marginRatio=0.10, maxSlope=None) -> tuple:
    """Localise le pic (maximum) du signal en excluant les bords.

    Les bords sont exclus par ``marginRatio`` pour ignorer les
    transitoires de dÃ©but/fin de balayage. Lorsque ``maxSlope`` est
    fourni, on Ã©limine les points dont la pente locale ``|dI/dV|``
    est supÃ©rieure au seuil : cela Ã©vite de confondre un flanc raide
    avec un vÃ©ritable sommet.

    ParamÃ¨tres:
        signalValues (numpy.ndarray): valeurs du signal (A).
        potentialValues (numpy.ndarray): potentiels associÃ©s (V),
            triÃ©s croissants.
        marginRatio (float): fraction de points Ã  ignorer de chaque
            cÃ´tÃ© (``0,10`` = 10 % des deux cÃ´tÃ©s).
        maxSlope (float | None): seuil absolu de pente. ``None``
            dÃ©sactive le filtre de pente.

    Retourne:
        tuple:
            - (float) potentiel du pic (V) ;
            - (float) amplitude du pic (A).
    """
    n = len(signalValues)
    margin = int(n * marginRatio)
    searchRegion = signalValues[margin:-margin]
    potentialsRegion = potentialValues[margin:-margin]

    if maxSlope is not None:
        slopes = np.gradient(searchRegion, potentialsRegion)
        validIndices = np.where(np.abs(slopes) < maxSlope)[0]
        if len(validIndices) == 0:
            # Aucun point ne satisfait le critÃ¨re de pente : repli sur la borne intÃ©rieure.
            return potentialValues[margin], signalValues[margin]
        bestIndex = validIndices[np.argmax(searchRegion[validIndices])]
        index = bestIndex + margin
    else:
        indexInRegion = np.argmax(searchRegion)
        index = indexInRegion + margin

    return potentialValues[index], signalValues[index]

def calculateSignalBaseLine(signalValues, potentialValues, xPeakVoltage, exclusionWidthRatio=0.03, lambdaFactor=1e3) -> tuple[np.ndarray, tuple[float, float]]:
    """Estime la ligne de base par l'algorithme asPLS.

    asPLS (*Asymmetric Penalized Least Squares Smoothing*, dÃ©rivÃ© de
    Whittaker) ajuste une courbe lisse Â« en dessous Â» du signal. Pour
    que la baseline ne suive pas le pic, on assigne un poids trÃ¨s
    faible (``0,001``) aux points situÃ©s dans une fenÃªtre centrÃ©e sur
    le pic : le solveur est alors libre d'ignorer ces points.

    Le paramÃ¨tre de rigiditÃ© ``lambda`` est mis Ã  l'Ã©chelle par ``nÂ²``
    afin d'Ãªtre indÃ©pendant de la densitÃ© d'Ã©chantillonnage : une
    mÃªme valeur de ``lambdaFactor`` produit un comportement similaire
    pour 200 ou 2 000 points.

    ParamÃ¨tres:
        signalValues (numpy.ndarray): signal lissÃ© (A).
        potentialValues (numpy.ndarray): potentiels associÃ©s (V).
        xPeakVoltage (float): potentiel du pic dÃ©tectÃ©, centre de la
            zone d'exclusion.
        exclusionWidthRatio (float): demi-largeur d'exclusion,
            exprimÃ©e en fraction de la plage totale de potentiel
            (``0,03`` = 3 %).
        lambdaFactor (float): multiplicateur de ``lambda``
            (``lambda = lambdaFactor * nÂ²``). Plus grand â baseline
            plus rigide.

    Retourne:
        tuple:
            - baselineValues (numpy.ndarray): ligne de base estimÃ©e.
            - (exclusion_min, exclusion_max) (tuple[float, float]):
              bornes de la zone d'exclusion, retournÃ©es pour tracÃ©
              ou debug.
    """
    n = len(signalValues)
    lam = lambdaFactor * (n ** 2)  # Mise Ã  l'Ã©chelle pour neutraliser l'effet du nombre de points.
    exclusionWidth = exclusionWidthRatio * (potentialValues[-1] - potentialValues[0])
    weights = np.ones_like(potentialValues)
    exclusion_min = xPeakVoltage - exclusionWidth
    exclusion_max = xPeakVoltage + exclusionWidth
    # PondÃ©ration quasi nulle dans la zone du pic : la baseline Â« passe sous Â» le pic.
    weights[(potentialValues > exclusion_min) & (potentialValues < exclusion_max)] = 0.001
    baselineValues, _ = aspls(signalValues, lam=lam, diff_order=2, weights=weights, tol=1e-2, max_iter=25)  # pyright: ignore[reportGeneralTypeIssues]
    return baselineValues, (exclusion_min, exclusion_max)

def plotSignalAnalysis(potentialValues, signalValues, signalSmoothed, baseline, signalCorrected, xCorrectedVoltage, yCorrectedCurrent, fileName, outputFolder) -> None:
    """Produit et enregistre un graphique rÃ©capitulatif au format PNG.

    Cinq tracÃ©s sont superposÃ©s afin de visualiser l'ensemble de la
    chaÃ®ne de traitement : signal brut semi-transparent, signal lissÃ©,
    baseline asPLS en tirets, signal corrigÃ©, et marqueur magenta Ã 
    la position du pic corrigÃ©. Le label du pic convertit le courant
    en mA pour lisibilitÃ©.

    Le PNG est sauvegardÃ© Ã  300 dpi pour une qualitÃ© d'impression.

    ParamÃ¨tres:
        potentialValues (numpy.ndarray): potentiels (V).
        signalValues (numpy.ndarray): signal brut (A, dÃ©jÃ  inversÃ©).
        signalSmoothed (numpy.ndarray): signal lissÃ© (A).
        baseline (numpy.ndarray): baseline asPLS (A).
        signalCorrected (numpy.ndarray): signal corrigÃ© (A).
        xCorrectedVoltage (float): potentiel du pic corrigÃ© (V).
        yCorrectedCurrent (float): amplitude du pic corrigÃ© (A).
        fileName (str): nom du fichier source (utilisÃ© pour le titre
            et le nom de sortie).
        outputFolder (str): dossier oÃ¹ enregistrer le PNG.
    """
    plt.figure(figsize=(10, 6))
    plt.plot(potentialValues, signalValues, label="Signal brut", alpha=0.5)
    plt.plot(potentialValues, signalSmoothed, label="Signal lissÃ©", linewidth=2)
    plt.plot(potentialValues, baseline, label="Baseline estimÃ©e (asPLS)", linestyle='--')
    plt.plot(potentialValues, signalCorrected, label="Signal corrigÃ©", linewidth=3)
    plt.plot(xCorrectedVoltage, yCorrectedCurrent, 'mo', label=f"Pic corrigÃ© Ã  {xCorrectedVoltage:.3f} V ({yCorrectedCurrent*1e3:.3f} mA)")
    plt.axvline(xCorrectedVoltage, color='magenta', linestyle=':', linewidth=1)
    plt.xlabel("Potentiel (V)")
    plt.ylabel("Courant (A)")
    plt.title(f"Correction de baseline : {fileName}")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    outputPath = os.path.join(outputFolder, fileName.replace(".txt", ".png"))
    plt.savefig(outputPath, dpi=300, bbox_inches='tight')
    plt.close()

def processFileWrapper(args):
    """Adapte ``processSignalFile`` Ã  ``multiprocessing.Pool.imap``.

    ``Pool.imap`` n'accepte qu'une fonction Ã  un seul argument ; ce
    wrapper dÃ©balle un tuple de paramÃ¨tres pour appeler la fonction
    rÃ©elle.

    ParamÃ¨tres:
        args (tuple): tuple Ã  dÃ©baller vers ``processSignalFile``.

    Retourne:
        dict | None: rÃ©sultat de ``processSignalFile``.
    """
    return processSignalFile(*args)

def processSignalFile(filePath, outputFolder, sep, decimal, export_choice) -> dict | None:
    """ExÃ©cute la chaÃ®ne complÃ¨te de traitement sur un fichier unique.

    Ãtapes : lecture â nettoyage â lissage â dÃ©tection de pic â
    estimation de baseline â correction â nouvelle dÃ©tection de pic
    â tracÃ© PNG â export optionnel â extraction des mÃ©tadonnÃ©es
    (base + Ã©lectrode via la regex ``(.+)_C(\\d{2})\\.txt``).

    Les erreurs sont capturÃ©es et retournÃ©es sous forme de
    dictionnaire ``{"error": ...}`` plutÃ´t que levÃ©es, afin que le
    pool de processus continue de traiter les autres fichiers.

    ParamÃ¨tres:
        filePath (str): chemin du fichier ``.txt`` Ã  traiter.
        outputFolder (str): dossier oÃ¹ Ã©crire PNG/CSV/XLSX.
        sep (str): sÃ©parateur de colonnes pour la lecture CSV.
        decimal (str): sÃ©parateur dÃ©cimal pour la lecture CSV.
        export_choice (int): ``0`` = pas d'export par fichier,
            ``1`` = export CSV, ``2`` = export XLSX.

    Retourne:
        dict: en cas de succÃ¨s, mapping vers une ligne du tableau
        rÃ©capitulatif (clÃ©s ``Base``, ``<electrode> - Tension (V)``,
        ``<electrode> - Courant (A)``, ``<electrode> - Charge (C)``).
        En cas d'Ã©chec, mapping ``{"error": "<message>"}``.
        ``None`` si la lecture renvoie un DataFrame vide.
    """
    try:
    # Extrait avant le try pour que fileName reste dÃ©fini si une exception
    # survient plus loin dans la chaÃ®ne.
        fileName = os.path.basename(filePath)
        dataFrame = readFile(filePath, sep=sep, decimal=decimal)
        if dataFrame is None:
            return None

        potentialValues, signalValues, cleaned_df = processData(dataFrame)
        signalSmoothed = smoothSignal(signalValues)
        # PremiÃ¨re dÃ©tection sur le signal lissÃ© : sert Ã  positionner la fenÃªtre d'exclusion asPLS.
        xPeakVoltage, _ = getPeakValue(signalSmoothed, potentialValues, marginRatio=0.10, maxSlope=500)
        baseline, _ = calculateSignalBaseLine(signalSmoothed, potentialValues, xPeakVoltage, exclusionWidthRatio=0.03, lambdaFactor=1e3)
        signalCorrected = signalSmoothed - baseline
        # Seconde dÃ©tection sur le signal corrigÃ© : c'est la valeur retenue pour le rÃ©capitulatif.
        xCorrectedVoltage, yCorrectedCurrent = getPeakValue(signalCorrected, potentialValues, marginRatio=0.10, maxSlope=500)
        plotSignalAnalysis(potentialValues, signalValues, signalSmoothed, baseline, signalCorrected, xCorrectedVoltage, yCorrectedCurrent, fileName, outputFolder)

        # Ajouter les colonnes calculÃ©es au DataFrame, pour l'Ã©ventuel export par fichier.
        cleaned_df["SignalLisse"] = signalSmoothed
        cleaned_df["SignalCorrigÃ©"] = signalCorrected

        if export_choice == 1:
            cleaned_df.to_csv(os.path.join(outputFolder, fileName.replace(".txt", ".csv")), index=False)
        elif export_choice == 2:
            cleaned_df.to_excel(os.path.join(outputFolder, fileName.replace(".txt", ".xlsx")), index=False)

        # Convention de nommage : <base>_C<NN>.txt â base + identifiant d'Ã©lectrode.
        # Si le nom ne suit pas la convention, on utilise le nom de fichier
        # comme base et l'Ã©lectrode reste vide.
        match = re.match(r"(.+)_C(\d{2})\.txt", fileName)
        baseName = match.group(1) if match else fileName
        electrode = f"C{match.group(2)}" if match else ""

        return {
            "Base": baseName,
            f"{electrode} - Tension (V)": xCorrectedVoltage,
            f"{electrode} - Courant (A)": yCorrectedCurrent,
            f"{electrode} - Charge (C)": "",  # Champ rempli ultÃ©rieurement par formule Excel.
        }

    except Exception as exception:  # pylint: disable=broad-exception-caught
        print(f"Erreur lors de la lecture de {filePath} : {exception}")
        return {"error": f"Erreur dans le fichier {os.path.basename(filePath)} : {str(exception)}"}

def main():
    """Point d'entrÃ©e : prÃ©pare le multiprocessing puis lance la GUI.

    ``freeze_support()`` est obligatoire sous Windows lorsque le
    programme est gelÃ© avec PyInstaller ou Ã©quivalent : sans cet
    appel, les processus enfants relancent le programme en boucle.
    """
    freeze_support()
    launch_gui()

def launch_gui():
    """Construit et exÃ©cute l'interface graphique Tkinter.

    Cette fonction encapsule toutes les variables et widgets Tkinter,
    ainsi que deux fonctions imbriquÃ©es (``select_folder`` et
    ``run_analysis``) qui forment la logique applicative dÃ©clenchÃ©e
    par les boutons.

    La fenÃªtre contient : un sÃ©lecteur de dossier, un cadre de
    paramÃ¨tres de lecture (sÃ©parateur de colonnes, sÃ©parateur
    dÃ©cimal, choix d'export), une barre de progression, un journal
    texte en temps rÃ©el et deux boutons d'action.

    La fonction bloque sur ``root.mainloop()`` jusqu'Ã  fermeture de
    la fenÃªtre.
    """
    def select_folder():
        """Ouvre un dialogue pour sÃ©lectionner le dossier d'entrÃ©e."""
        path = filedialog.askdirectory(title="SÃ©lectionnez le dossier contenant les fichiers .txt")
        if path:
            folder_path.set(path)

    def run_analysis():
        """Pipeline global : prÃ©paration, parallÃ©lisation, agrÃ©gation Excel.

        Ãtapes principales :
            1. validation du dossier d'entrÃ©e ;
            2. rÃ©solution des sÃ©parateurs (libellÃ©s GUI â caractÃ¨res) ;
            3. crÃ©ation/nettoyage du dossier ``<entrÃ©e> (results)`` ;
            4. collecte de tous les ``.txt`` du dossier ;
            5. traitement parallÃ¨le via ``multiprocessing.Pool`` avec
               ``imap`` afin d'afficher les logs au fil de l'eau ;
            6. construction du DataFrame rÃ©capitulatif, regroupement
               par ``Base`` et injection d'une frÃ©quence Ã  50 Hz ;
            7. rÃ©Ã©criture du fichier Excel via ``openpyxl`` pour
               remplacer la colonne ``Charge (C)`` par une formule
               ``=Courant/FrÃ©q`` (calculÃ©e dynamiquement par Excel).
        """
        export_choice = export_option.get()
        #export_csv = export_choice == 1
        #export_excel = export_choice == 2

        log_box.config(state="normal")
        log_box.delete("1.0", "end")
        log_box.config(state="disabled")
        inputFolder = folder_path.get()
        if not inputFolder or not os.path.isdir(inputFolder):
            messagebox.showerror("Erreur", "Veuillez sÃ©lectionner un dossier valide.")
            return

        # Conversion des libellÃ©s affichÃ©s dans la GUI vers les caractÃ¨res rÃ©els.
        sep_label = sep_var.get()
        sep_map = {"Tabulation": "\t", "Virgule": ",", "Point-virgule": ";", "Espace": " "}
        sep = sep_map.get(sep_label, "\t")
        decimal_label = decimal_var.get()
        decimal_map = {"Point": ".", "Virgule": ","}
        decimal = decimal_map.get(decimal_label, ".")

        # Dossier de sortie : crÃ©Ã© Ã  cÃ´tÃ© du dossier d'entrÃ©e, suffixÃ© Â« (results) Â»
        # (p. ex. Â« campagne_mars (results) Â»).
        folderName = os.path.basename(os.path.normpath(inputFolder))
        outputFolder = os.path.join(os.path.dirname(inputFolder), folderName + " (results)")
        os.makedirs(outputFolder, exist_ok=True)

        # Nettoyage des artefacts d'une exÃ©cution prÃ©cÃ©dente pour Ã©viter toute
        # confusion entre une campagne antÃ©rieure et la campagne en cours.
        log_box.config(state="normal")
        log_box.insert("end", "Nettoyage du dossier de sortie...\n")
        log_box.config(state="disabled")
        for file in glob.glob(os.path.join(outputFolder, "*")):
            if file.endswith((".png", ".csv", ".xlsx")):
                os.remove(file)

        filePaths = sorted(glob.glob(os.path.join(inputFolder, "*.txt")))
        fileProcessingArgs = [(filePath, outputFolder, sep, decimal, export_choice) for filePath in filePaths]

        results = []
        start_time = time.time()

        progress_bar["maximum"] = len(filePaths)
        progress_bar["value"] = 0

        def iter_results():
            """ItÃ¨re sur les rÃ©sultats selon le mode sÃ©lectionnÃ© (parallÃ¨le ou sÃ©quentiel).

            Le ``with Pool(...)`` est encapsulÃ© ici afin que son scope se ferme
            proprement quand le gÃ©nÃ©rateur est Ã©puisÃ©, sans dupliquer le corps
            de boucle appelant (log_box, progress_bar, gestion d'erreur).
            """
            if multi_thread_option.get() == 1:
                # Mode parallÃ¨le : un processus par cÅur ; imap permet d'itÃ©rer les
                # rÃ©sultats au fur et Ã  mesure pour rafraÃ®chir logs + barre de progression.
                with Pool(processes=cpu_count()) as pool:
                    yield from pool.imap(processFileWrapper, fileProcessingArgs)
            else:
                # Mode sÃ©quentiel : traitement fichier par fichier dans le processus principal.
                for args in fileProcessingArgs:
                    yield processFileWrapper(args)

        for i, (filePath, result) in enumerate(zip(filePaths, iter_results())):
            log_box.config(state="normal")
            if result:
                if "error" in result:
                    log_box.insert("end", f"Erreur : {result['error']}\n", ("error",))
                else:
                    results.append(result)
                    log_box.insert("end", f"Traitement : {os.path.basename(filePath)}\n")
            else:
                log_box.insert("end", f"Fichier ignorÃ© ou invalide : {os.path.basename(filePath)}\n")

            log_box.update_idletasks()
            log_box.see("end")
            log_box.tag_config("error", foreground="red")
            log_box.config(state="disabled")
            progress_bar["value"] = i + 1
            root.update_idletasks()

        if results:
            # AgrÃ©gation : une ligne par base, fusion des colonnes par Ã©lectrode.
            # ``groupby().first()`` regroupe toutes les Ã©lectrodes d'une mÃªme
            # base sur une seule ligne, chaque Ã©lectrode occupant ses propres
            # colonnes Tension/Courant/Charge.
            df = pd.DataFrame(results)
            df = df.groupby("Base").first().reset_index()
            # FrÃ©quence SWV par dÃ©faut (50 Hz) â dÃ©nominateur de la formule de charge ;
            # modifiable manuellement dans Excel aprÃ¨s gÃ©nÃ©ration.
            df.insert(1, 'FrÃ©q (Hz)', 50.0)
            excel_path = os.path.join(outputFolder, folderName + ".xlsx")
            df.to_excel(excel_path, index=False)

            # Post-traitement avec openpyxl : injection de formules Excel pour la charge,
            # afin que l'utilisateur puisse modifier la frÃ©quence et voir la charge
            # recalculÃ©e automatiquement dans Excel.
            wb = load_workbook(excel_path)
            ws = wb.active
            # ``wb.active`` est typÃ© Optional mais renvoie toujours une feuille
            # pour un classeur fraÃ®chement Ã©crit : assertion pour le type-checker.
            assert ws is not None
            # Les en-tÃªtes de colonnes sont toujours des chaÃ®nes dans ce
            # classeur ; ``cast`` l'indique Ã  Pyright sans transformation au
            # runtime.
            header = [cast(str, cell.value) for cell in ws[1]]
            freq_col_letter = get_column_letter(header.index('FrÃ©q (Hz)') + 1)

            for col_index, col_name in enumerate(header):
                if col_name.endswith("- Courant (A)"):
                    elec = col_name.split(" - ")[0]
                    charge_col = f"{elec} - Charge (C)"
                    if charge_col in header:
                        charge_col_index = header.index(charge_col) + 1
                        current_col_letter = get_column_letter(col_index + 1)
                        for row in range(2, ws.max_row + 1):
                            formula = f"={current_col_letter}{row}/{freq_col_letter}{row}"
                            ws.cell(row=row, column=charge_col_index, value=formula)

            wb.save(excel_path)
            log_box.config(state="normal")
            duration = time.time() - start_time
            summary = f"\nTraitement terminÃ© avec succÃ¨s.\nFichiers traitÃ©s : {len(results)} / {len(filePaths)}\nTemps Ã©coulÃ© : {duration:.2f} secondes.\n\n"
            log_box.insert("end", summary)
            log_box.update_idletasks()
            log_box.see("end")
            log_box.config(state="disabled")
            messagebox.showinfo("SuccÃ¨s", "Traitement terminÃ© avec succÃ¨s.")
            result_button.config(state="normal")

    # --- Construction de la fenÃªtre principale -----------------------------
    root = Tk()
    root.resizable(True, True)

    root.title("Analyse de fichiers SWV")
    root.geometry("700x400")
    root.minsize(600, 400)

    # Variables Tkinter partagÃ©es entre les widgets et les callbacks.
    folder_path = StringVar()
    sep_options = ["Tabulation", "Virgule", "Point-virgule", "Espace"]
    decimal_options = ["Point", "Virgule"]

    sep_var = StringVar(value="Tabulation")
    decimal_var = StringVar(value="Point")
    export_option = IntVar(value=0)
    multi_thread_option = IntVar(value=1)  # 1 = activÃ© par dÃ©faut (comportement historique).

    # Cadre principal : une grille extensible qui hÃ©berge les 5 blocs visuels.
    main_frame = Frame(root, padx=10, pady=10)
    main_frame.grid(row=0, column=0, sticky="nsew")
    main_frame.grid_columnconfigure(1, weight=1)
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Ligne 0 : sÃ©lection du dossier d'entrÃ©e.
    Label(main_frame, text="Dossier d'entrÃ©e :").grid(row=0, column=0, sticky="w")
    Label(main_frame, textvariable=folder_path, relief="sunken", anchor="w", width=50).grid(row=0, column=1, padx=5, sticky="ew")
    Button(main_frame, text="Parcourir", command=select_folder).grid(row=0, column=2, padx=5)

    # Ligne 1 : paramÃ¨tres de lecture (sÃ©parateurs + choix d'export).
    settings_frame = ttk.LabelFrame(main_frame, text="ParamÃ¨tres de lecture")
    settings_frame.grid(row=1, column=0, columnspan=3, pady=(10, 5), sticky="ew")

    Label(settings_frame, text="SÃ©parateur de colonnes :").grid(row=0, column=0, sticky="w")
    sep_radio_frame = Frame(settings_frame)
    sep_radio_frame.grid(row=0, column=1, columnspan=4, sticky="w")
    for i, txt in enumerate(sep_options):
        ttk.Radiobutton(sep_radio_frame, text=txt, variable=sep_var, value=txt).grid(row=0, column=i, sticky="w", padx=(0, 10))

    Label(settings_frame, text="SÃ©parateur dÃ©cimal :").grid(row=1, column=0, sticky="w")
    dec_radio_frame = Frame(settings_frame)
    dec_radio_frame.grid(row=1, column=1, columnspan=4, sticky="w")
    for i, txt in enumerate(decimal_options):
        ttk.Radiobutton(dec_radio_frame, text=txt, variable=decimal_var, value=txt).grid(row=0, column=i, sticky="w", padx=(0, 10))

    Label(settings_frame, text="Export des fichiers :").grid(row=2, column=0, sticky="w", pady=(5, 0))
    export_radio_frame = Frame(settings_frame)
    export_radio_frame.grid(row=2, column=1, columnspan=4, sticky="w")
    Radiobutton(export_radio_frame, text="Ne pas exporter", variable=export_option, value=0).pack(side="left", padx=(0, 10))
    Radiobutton(export_radio_frame, text="Exporter au format .CSV", variable=export_option, value=1).pack(side="left", padx=(0, 10))
    Radiobutton(export_radio_frame, text="Exporter au format Excel", variable=export_option, value=2).pack(side="left")

    Label(settings_frame, text="Traitement parallÃ¨le :").grid(row=3, column=0, sticky="w", pady=(5, 0))
    multi_thread_radio_frame = Frame(settings_frame)
    multi_thread_radio_frame.grid(row=3, column=1, columnspan=4, sticky="w")
    Radiobutton(multi_thread_radio_frame, text="Activer le multi-thread (un processus par cÅur)", variable=multi_thread_option, value=1).pack(side="left", padx=(0, 10))
    Radiobutton(multi_thread_radio_frame, text="DÃ©sactiver (traitement sÃ©quentiel)", variable=multi_thread_option, value=0).pack(side="left")

    # Ligne 2 : barre de progression (mode Â« determinate Â», maximum = nb fichiers).
    progress_frame = ttk.LabelFrame(main_frame, text="Progression du traitement")
    progress_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=2, pady=(5, 5))
    progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate")
    progress_bar.pack(fill="x", padx=5, pady=5)

    # Ligne 3 : journal de traitement (widget Text en lecture seule, extensible).
    log_frame = ttk.LabelFrame(main_frame, text="Journal de traitement")
    log_frame.grid(row=3, column=0, columnspan=3, sticky="nsew", padx=2, pady=(0, 5))
    main_frame.grid_rowconfigure(3, weight=1)
    log_box = Text(log_frame, relief="sunken", wrap="word", height=10, bg="white")
    log_box.pack(expand=True, fill="both", padx=5, pady=5)
    log_box.config(state="disabled")

    # Ligne 4 : boutons d'action (lancement + ouverture du dossier de rÃ©sultats).
    action_frame = Frame(main_frame)
    action_frame.grid(row=4, column=0, columnspan=3, sticky="ew")
    Button(action_frame, text="Lancer l'analyse", command=run_analysis).pack(side="right", padx=5, pady=5)
    result_button = Button(action_frame, text="Ouvrir le dossier de rÃ©sultats", state="disabled", command=lambda: open_folder(folder_path.get() + " (results)"))
    result_button.pack(side="right", padx=5, pady=5)

    root.mainloop()

if __name__ == '__main__':
    main()
